"""
export_utils.py — Export avancé en PDF, Excel et JSON
Formats multiples avec validation et gestion des erreurs
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from datetime import datetime
from typing import Optional, TYPE_CHECKING
import csv
import os
import json

if TYPE_CHECKING:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore

class ExportManager:
    """Gestion des exports PDF, Excel, CSV et JSON"""
    
    @staticmethod
    def export_students_pdf(students: list, filename: str = "etudiants.pdf") -> Optional[str]:
        """Exporte la liste des étudiants en PDF avec style moderne"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=20, bottomMargin=20)
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre avec style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#D4AF37'),  # Doré
                spaceAfter=30,
                alignment=1  # Center
            )
            elements.append(Paragraph("Liste des Étudiants - DanProject", title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Date génération
            date_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Table
            data = [["Nom", "Prénom", "Classe", "Email"]]
            for student in students:
                data.append([
                    student.get('lastname', '-'),
                    student.get('firstname', '-'),
                    student.get('class_name', '-'),
                    student.get('email', '-')
                ])
            
            table = Table(data, colWidths=[2*inch, 2*inch, 1.5*inch, 2.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),  # Bleu
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FC')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            return filename
        except Exception as e:
            print(f"Erreur export PDF: {e}")
            return None
    
    @staticmethod
    def export_absences_pdf(absences: list, filename: str = "absences.pdf") -> Optional[str]:
        """Exporte la liste des absences en PDF"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=20, bottomMargin=20)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#D4AF37'),  # Doré
                spaceAfter=30,
                alignment=1
            )
            elements.append(Paragraph("Rapport des Absences - DanProject", title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            date_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Table
            data = [["Étudiant", "Date", "Statut", "Justification"]]
            for absence in absences:
                justif = absence.get('justification', '')
                if len(justif) > 40:
                    justif = justif[:40] + '...'
                data.append([
                    absence.get('student_name', '-'),
                    absence.get('date_absence', '-'),
                    absence.get('status', '-'),
                    justif
                ])
            
            table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1.2*inch, 2.3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FC')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            return filename
        except Exception as e:
            print(f"Erreur export PDF: {e}")
            return None
    
    @staticmethod
    def export_to_csv(data: list, columns: list, filename: str) -> Optional[str]:
        """Exporte les données en CSV"""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=columns)
                writer.writeheader()
                writer.writerows(data)
            return filename
        except Exception as e:
            print(f"Erreur export CSV: {e}")
            return None
    
    @staticmethod
    def export_to_json(data: list, filename: str) -> Optional[str]:
        """Exporte les données en JSON"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return filename
        except Exception as e:
            print(f"Erreur export JSON: {e}")
            return None
    
    @staticmethod
    def export_to_excel(data: list, columns: list, filename: str) -> Optional[str]:
        """Exporte les données en Excel (si openpyxl est installé)"""
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
            
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Impossible de créer une feuille de calcul")
            
            ws.title = "Données"
            
            # En-têtes avec style
            header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col)
                if cell is not None:
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
            
            # Données avec style
            data_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            for row, record in enumerate(data, 2):
                for col, field in enumerate(columns, 1):
                    cell = ws.cell(row=row, column=col)
                    if cell is not None:
                        cell.value = record.get(field, '')
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                        
                        # Couleur alternée
                        if row % 2 == 0:
                            cell.fill = data_fill
            
            # Ajuster la largeur des colonnes
            try:
                for col in ws.columns:
                    if not col:
                        continue
                    max_length = 0
                    # Récupérer le column_letter du premier cell non-MergedCell
                    column_letter = None
                    for cell in col:
                        try:
                            if hasattr(cell, 'column_letter'):
                                column_letter = cell.column_letter  # type: ignore
                                break
                            elif cell.coordinate:
                                column_letter = cell.coordinate[0]
                                break
                        except:
                            pass
                    
                    if not column_letter:
                        continue
                    
                    for cell in col:
                        try:
                            if cell and cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            except:
                pass
            
            wb.save(filename)
            return filename
        except ImportError:
            # Fallback vers CSV si openpyxl n'est pas installé
            print("openpyxl non installé, export en CSV à la place")
            return ExportManager.export_to_csv(data, columns, filename.replace('.xlsx', '.csv'))
        except Exception as e:
            print(f"Erreur export Excel: {e}")
            return None
    
    @staticmethod
    def export_report(title: str, sections: dict, filename: str = "rapport.pdf") -> Optional[str]:
        """Exporte un rapport complet avec plusieurs sections"""
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.HexColor('#D4AF37'),
                spaceAfter=20,
                alignment=1
            )
            
            section_style = ParagraphStyle(
                'Section',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#1E40AF'),
                spaceAfter=10,
                spaceBefore=10
            )
            
            elements.append(Paragraph(title, title_style))
            elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", 
                                styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            for section_name, section_content in sections.items():
                elements.append(Paragraph(section_name, section_style))
                if isinstance(section_content, str):
                    elements.append(Paragraph(section_content, styles['Normal']))
                elements.append(Spacer(1, 0.2*inch))
            
            doc.build(elements)
            return filename
        except Exception as e:
            print(f"Erreur export rapport: {e}")
            return None
